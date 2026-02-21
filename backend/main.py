from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pathlib import Path
import csv
import io
import os
import time
import json
import zipfile
from datetime import date, datetime
from typing import Optional, List, Dict, Any

from dotenv import load_dotenv
from supabase import create_client, Client

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

from groq import Groq

from PIL import Image
import pytesseract


# --------------------
# Config / Setup
# --------------------
ENV_PATH = Path(__file__).parent / ".env"
load_dotenv(dotenv_path=ENV_PATH)

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")

TESSERACT_CMD = os.getenv("TESSERACT_CMD")
if TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

supabase: Optional[Client] = None
if SUPABASE_URL and SUPABASE_SERVICE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

groq_client: Optional[Groq] = None
if GROQ_API_KEY:
    groq_client = Groq(api_key=GROQ_API_KEY)

app = FastAPI(title="CRM Deals Consolidation", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOADS_DIR = Path(__file__).parent / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)


# --------------------
# Helpers
# --------------------
def require_supabase():
    if supabase is None:
        raise HTTPException(status_code=500, detail="Supabase not configured. Check backend/.env")


def require_groq():
    if groq_client is None:
        raise HTTPException(status_code=500, detail="Groq not configured. Set GROQ_API_KEY in backend/.env")


def to_int(value):
    try:
        s = str(value).strip()
        if s == "":
            return None
        return int(float(s))
    except Exception:
        return None


def to_date(value):
    try:
        if value is None:
            return None
        s = str(value).strip()
        if s == "":
            return None
        return date.fromisoformat(s)
    except Exception:
        return None


def normalize_deal(row: dict, upload_timestamp: str, processing_status: str) -> dict:
    deal = {
        "deal_id": str(row.get("deal_id", "")).strip(),
        "client_name": str(row.get("client_name", "")).strip(),
        "deal_value": to_int(row.get("deal_value", "")),
        "stage": str(row.get("stage", "")).strip(),
        "closing_probability": to_int(row.get("closing_probability", "")),
        "owner": str(row.get("owner", "")).strip(),
        "expected_close_date": None,
        "upload_timestamp": upload_timestamp,
        "processing_status": processing_status,
    }

    d = row.get("expected_close_date")
    if isinstance(d, datetime):
        deal["expected_close_date"] = d.date().isoformat()
    elif isinstance(d, date):
        deal["expected_close_date"] = d.isoformat()
    else:
        parsed = to_date(d)
        if parsed:
            deal["expected_close_date"] = parsed.isoformat()

    return deal


def save_upload(file: UploadFile, fallback_name: str) -> tuple[str, Path]:
    safe_name = (file.filename or fallback_name).replace(" ", "_")
    saved_name = f"{int(time.time())}_{safe_name}"
    saved_path = UPLOADS_DIR / saved_name
    return saved_name, saved_path


def file_kind_from_name(name: str) -> str:
    n = (name or "").lower()
    if n.endswith(".csv"):
        return "csv"
    if n.endswith(".xlsx"):
        return "xlsx"
    if n.endswith(".pdf"):
        return "pdf"
    if n.endswith(".png") or n.endswith(".jpg") or n.endswith(".jpeg"):
        return "image"
    return "unknown"


def is_supported(name: str) -> bool:
    return file_kind_from_name(name) in {"csv", "xlsx", "pdf", "image"}


def safe_extract_zip(z: zipfile.ZipFile, dest_dir: Path) -> List[Path]:
    """
    Extract zip safely (avoid path traversal). Returns extracted file paths.
    """
    extracted: List[Path] = []
    dest_dir.mkdir(parents=True, exist_ok=True)

    for member in z.infolist():
        if member.is_dir():
            continue

        # prevent zip slip
        member_name = member.filename.replace("\\", "/")
        if member_name.startswith("/") or ".." in member_name.split("/"):
            continue

        out_path = dest_dir / member_name
        out_path.parent.mkdir(parents=True, exist_ok=True)

        with z.open(member) as src, open(out_path, "wb") as dst:
            dst.write(src.read())

        extracted.append(out_path)

    return extracted


# --------------------
# Processing functions (from file paths)
# --------------------
def process_csv_path(path: Path) -> Dict[str, Any]:
    require_supabase()

    upload_timestamp = datetime.utcnow().isoformat() + "Z"
    processing_status = "inserted"

    text = path.read_text(encoding="utf-8-sig", errors="replace")

    cleaned_lines = []
    for line in text.splitlines():
        line = line.strip()
        if line.startswith('"') and line.endswith('"'):
            line = line[1:-1]
        cleaned_lines.append(line)
    cleaned_text = "\n".join(cleaned_lines)

    reader = csv.DictReader(io.StringIO(cleaned_text), delimiter=",")

    headers = reader.fieldnames or []
    preview = []
    rows_to_insert = []
    total_rows = 0

    for i, row in enumerate(reader):
        total_rows += 1
        deal = normalize_deal(row, upload_timestamp, processing_status)
        rows_to_insert.append(deal)
        if i < 5:
            preview.append(deal)

    if total_rows == 0:
        return {"file_type": "csv", "total_rows": 0, "inserted_rows": 0, "preview": []}

    result = supabase.table("deals").insert(rows_to_insert).execute()
    inserted_rows = len(result.data) if result.data else 0

    return {
        "file_type": "csv",
        "headers": headers,
        "total_rows": total_rows,
        "inserted_rows": inserted_rows,
        "preview": preview,
    }


def process_xlsx_path(path: Path) -> Dict[str, Any]:
    require_supabase()

    upload_timestamp = datetime.utcnow().isoformat() + "Z"
    processing_status = "inserted"

    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    sheet_used = ws.title
    rows = list(ws.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        return {
            "file_type": "xlsx",
            "sheet_used": sheet_used,
            "total_rows": 0,
            "inserted_rows": 0,
            "preview": [],
        }

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    preview = []
    rows_to_insert = []
    total_rows = 0

    for values in rows[1:]:
        row_dict = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            row_dict[h] = values[idx] if idx < len(values) else None

        if all((v is None or str(v).strip() == "") for v in row_dict.values()):
            continue

        total_rows += 1
        deal = normalize_deal(row_dict, upload_timestamp, processing_status)
        rows_to_insert.append(deal)
        if total_rows <= 5:
            preview.append(deal)

    if total_rows == 0:
        return {
            "file_type": "xlsx",
            "sheet_used": sheet_used,
            "total_rows": 0,
            "inserted_rows": 0,
            "preview": [],
        }

    result = supabase.table("deals").insert(rows_to_insert).execute()
    inserted_rows = len(result.data) if result.data else 0

    return {
        "file_type": "xlsx",
        "sheet_used": sheet_used,
        "headers": headers,
        "total_rows": total_rows,
        "inserted_rows": inserted_rows,
        "preview": preview,
    }


def process_pdf_path(path: Path) -> Dict[str, Any]:
    require_supabase()

    upload_timestamp = datetime.utcnow().isoformat() + "Z"
    processing_status = "extracted"

    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t.strip():
            parts.append(t)
    raw_text = "\n\n".join(parts).strip()

    if not raw_text:
        processing_status = "no_text_found"

    doc_row = {
        "source_file": path.name,
        "file_type": "pdf",
        "raw_text": raw_text,
        "upload_timestamp": upload_timestamp,
        "processing_status": processing_status,
    }

    result = supabase.table("documents").insert(doc_row).execute()
    inserted_id = result.data[0]["id"] if result.data else None

    return {
        "file_type": "pdf",
        "document_id": inserted_id,
        "processing_status": processing_status,
        "text_length": len(raw_text),
        "preview": raw_text[:1200] if raw_text else "",
    }


def process_image_path(path: Path) -> Dict[str, Any]:
    require_supabase()

    upload_timestamp = datetime.utcnow().isoformat() + "Z"
    processing_status = "ocr_extracted"

    img = Image.open(path)
    raw_text = (pytesseract.image_to_string(img) or "").strip()

    if not raw_text:
        processing_status = "no_text_found"

    doc_row = {
        "source_file": path.name,
        "file_type": "image",
        "raw_text": raw_text,
        "upload_timestamp": upload_timestamp,
        "processing_status": processing_status,
    }

    result = supabase.table("documents").insert(doc_row).execute()
    inserted_id = result.data[0]["id"] if result.data else None

    return {
        "file_type": "image",
        "document_id": inserted_id,
        "processing_status": processing_status,
        "text_length": len(raw_text),
        "preview": raw_text[:1200] if raw_text else "",
    }


def process_any_path(path: Path) -> Dict[str, Any]:
    kind = file_kind_from_name(path.name)
    if kind == "csv":
        return process_csv_path(path)
    if kind == "xlsx":
        return process_xlsx_path(path)
    if kind == "pdf":
        return process_pdf_path(path)
    if kind == "image":
        return process_image_path(path)
    return {"file_type": "unknown", "error": "Unsupported file"}


# --------------------
# Basic routes
# --------------------
@app.get("/health")
def health_check():
    return {"status": "ok"}


@app.get("/")
def root():
    return {"app": "crm-deals-consolidator", "status": "running"}


# --------------------
# Existing single-file endpoints
# --------------------
@app.post("/upload/csv")
async def upload_csv(file: UploadFile = File(...)):
    require_supabase()
    if not (file.filename or "").lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a .csv file")

    saved_name, saved_path = save_upload(file, "upload.csv")
    saved_path.write_bytes(await file.read())

    out = process_csv_path(saved_path)
    out["saved_as"] = saved_name
    return out


@app.post("/upload/xlsx")
async def upload_xlsx(file: UploadFile = File(...)):
    require_supabase()
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx file")

    saved_name, saved_path = save_upload(file, "upload.xlsx")
    saved_path.write_bytes(await file.read())

    out = process_xlsx_path(saved_path)
    out["saved_as"] = saved_name
    return out


@app.post("/upload/pdf")
async def upload_pdf(file: UploadFile = File(...)):
    require_supabase()
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Please upload a .pdf file")

    saved_name, saved_path = save_upload(file, "upload.pdf")
    saved_path.write_bytes(await file.read())

    out = process_pdf_path(saved_path)
    out["saved_as"] = saved_name
    return out


@app.post("/upload/image")
async def upload_image(file: UploadFile = File(...)):
    require_supabase()
    n = (file.filename or "").lower()
    if not (n.endswith(".png") or n.endswith(".jpg") or n.endswith(".jpeg")):
        raise HTTPException(status_code=400, detail="Please upload a .png, .jpg, or .jpeg image")

    saved_name, saved_path = save_upload(file, "upload.png")
    saved_path.write_bytes(await file.read())

    out = process_image_path(saved_path)
    out["saved_as"] = saved_name
    return out


# --------------------
# ZIP: upload & list contents (manual processing)
# --------------------
@app.post("/upload/zip")
async def upload_zip(file: UploadFile = File(...)):
    require_supabase()

    if not (file.filename or "").lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload a .zip file")

    saved_name, saved_path = save_upload(file, "upload.zip")
    saved_path.write_bytes(await file.read())

    zip_id = f"zip_{int(time.time())}"
    extract_dir = UPLOADS_DIR / zip_id

    try:
        with zipfile.ZipFile(saved_path, "r") as z:
            extracted_paths = safe_extract_zip(z, extract_dir)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not extract zip: {e}")

    # only show supported files
    files = []
    for p in extracted_paths:
        if is_supported(p.name):
            rel = str(p.relative_to(extract_dir)).replace("\\", "/")
            files.append(
                {
                    "name": rel,
                    "file_type": file_kind_from_name(p.name),
                    "size": p.stat().st_size if p.exists() else 0,
                }
            )

    # stable ordering for UI
    files.sort(key=lambda x: x["name"])

    return {
        "file_type": "zip",
        "saved_as": saved_name,
        "zip_id": zip_id,
        "extracted_count": len(extracted_paths),
        "supported_files": files,
    }


@app.post("/zip/{zip_id}/process")
def process_zip_files(zip_id: str, payload: Dict[str, Any]):
    """
    payload = { "files": ["path/in/zip/file1.pdf", "file2.csv"] }
    """
    require_supabase()

    files = payload.get("files", [])
    if not isinstance(files, list) or not files:
        raise HTTPException(status_code=400, detail="Provide a JSON body like: {\"files\": [\"file1.pdf\"]}")

    extract_dir = UPLOADS_DIR / zip_id
    if not extract_dir.exists():
        raise HTTPException(status_code=404, detail="zip_id not found")

    results = []
    deals_inserted_total = 0
    documents_created = []

    for rel in files:
        try:
            rel_clean = str(rel).replace("\\", "/").lstrip("/")
            target = extract_dir / rel_clean
            if not target.exists() or not target.is_file():
                results.append({"name": rel, "status": "missing"})
                continue

            if not is_supported(target.name):
                results.append({"name": rel, "status": "skipped", "reason": "unsupported"})
                continue

            out = process_any_path(target)
            out["name"] = rel
            out["status"] = "processed"
            results.append(out)

            if out.get("file_type") in {"csv", "xlsx"}:
                deals_inserted_total += int(out.get("inserted_rows") or 0)

            if out.get("file_type") in {"pdf", "image"} and out.get("document_id"):
                documents_created.append(
                    {"name": rel, "file_type": out.get("file_type"), "document_id": out.get("document_id")}
                )

        except Exception as e:
            results.append({"name": rel, "status": "error", "error": str(e)})

    return {
        "zip_id": zip_id,
        "processed_files": len(results),
        "deals_inserted_total": deals_inserted_total,
        "documents_created": documents_created,
        "results": results,
    }


# --------------------
# LLM Structuring
# --------------------
@app.post("/documents/{document_id}/structure")
def structure_document(document_id: str):
    require_supabase()
    require_groq()

    doc_resp = supabase.table("documents").select("*").eq("id", document_id).execute()
    rows = doc_resp.data or []
    if not rows:
        raise HTTPException(status_code=404, detail="Document not found")

    doc = rows[0]
    raw_text = (doc.get("raw_text") or "").strip()
    if not raw_text:
        try:
            supabase.table("documents").update({"processing_status": "no_text_found"}).eq("id", document_id).execute()
        except Exception:
            pass
        raise HTTPException(status_code=400, detail="Document has no text to structure")

    text_for_llm = raw_text[:12000]

    system_prompt = """
You extract CRM deals from text and return ONLY JSON.
Return a single JSON object with this shape:

{
  "deals": [
    {
      "deal_id": "string",
      "client_name": "string",
      "deal_value": number_or_null,
      "stage": "string",
      "closing_probability": number_or_null,
      "owner": "string",
      "expected_close_date": "YYYY-MM-DD_or_empty"
    }
  ]
}

Rules:
- Output must be valid JSON (no markdown, no commentary).
- If a field is missing, use empty string for text fields, null for numbers.
- expected_close_date must be "YYYY-MM-DD" if known, otherwise "".
- Return at most 50 deals.
""".strip()

    user_prompt = f"Text to extract from:\n\n{text_for_llm}"

    try:
        completion = groq_client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            temperature=0.1,
        )
        content = completion.choices[0].message.content or "{}"
    except Exception as e:
        try:
            supabase.table("documents").update({"processing_status": "llm_failed"}).eq("id", document_id).execute()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"Groq request failed: {e}")

    try:
        data = json.loads(content)
    except Exception as e:
        try:
            supabase.table("documents").update({"processing_status": "bad_json"}).eq("id", document_id).execute()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"LLM returned invalid JSON: {e}")

    deals = data.get("deals", [])
    if not isinstance(deals, list):
        raise HTTPException(status_code=500, detail="LLM JSON missing 'deals' list")

    upload_timestamp = datetime.utcnow().isoformat() + "Z"
    processing_status = "structured"

    normalized = []
    for d in deals[:50]:
        if isinstance(d, dict):
            normalized.append(normalize_deal(d, upload_timestamp, processing_status))

    if not normalized:
        try:
            supabase.table("documents").update({"processing_status": "structured_no_deals"}).eq("id", document_id).execute()
        except Exception:
            pass
        return {"document_id": document_id, "structured_deals": 0, "inserted_rows": 0, "preview": []}

    try:
        ins = supabase.table("deals").insert(normalized).execute()
        inserted_rows = len(ins.data) if ins.data else 0
    except Exception as e:
        try:
            supabase.table("documents").update({"processing_status": "db_insert_failed"}).eq("id", document_id).execute()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"Supabase insert failed: {e}")

    try:
        supabase.table("documents").update({"processing_status": "structured"}).eq("id", document_id).execute()
    except Exception:
        pass

    return {
        "document_id": document_id,
        "model": GROQ_MODEL,
        "structured_deals": len(normalized),
        "inserted_rows": inserted_rows,
        "preview": normalized[:5],
    }


# --------------------
# Export deals as Excel
# --------------------
@app.get("/export")
def export_deals_to_excel():
    require_supabase()

    all_rows = []
    start = 0
    step = 1000

    while True:
        resp = supabase.table("deals").select("*").range(start, start + step - 1).execute()
        chunk = resp.data or []
        all_rows.extend(chunk)
        if len(chunk) < step:
            break
        start += step

    wb = Workbook()
    ws = wb.active
    ws.title = "deals"

    columns = [
        "deal_id",
        "client_name",
        "deal_value",
        "stage",
        "closing_probability",
        "owner",
        "expected_close_date",
        "upload_timestamp",
        "processing_status",
    ]
    ws.append(columns)

    for r in all_rows:
        ws.append([
            r.get("deal_id"),
            r.get("client_name"),
            r.get("deal_value"),
            r.get("stage"),
            r.get("closing_probability"),
            r.get("owner"),
            r.get("expected_close_date"),
            r.get("upload_timestamp"),
            r.get("processing_status"),
        ])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"crm_deals_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )